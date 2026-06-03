#!/usr/bin/env python3
"""Import MoneyHunter into a public static learning hub.

The importer keeps the source material public and complete by default:
- copy every non-.git, non-.DS_Store file into raw/
- extract every URL from Markdown/text files
- generate a browsable/searchable static site
- generate docs that explain how to study the archive

Only high-confidence credential material blocks the import.
"""

from __future__ import annotations

import argparse
import hashlib
import html
import json
import os
import re
import shutil
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, quote, unquote, urlparse
import xml.etree.ElementTree as ET


DEFAULT_SOURCE = Path(
    "/Users/siuserxiaowei/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_276exkqyuyd422_20a2/msg/file/2026-02/MoneyHunter"
)
ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "raw"
SITE_DIR = ROOT / "site"
DOCS_DIR = ROOT / "docs"
DATA_DIR = ROOT / "data"

URL_RE = re.compile(r"https?://[^\s<>'\"`\\]+", re.IGNORECASE)
WWW_RE = re.compile(r"(?<![@/\w])(www\.[A-Za-z0-9.-]+\.[A-Za-z]{2,}(?:/[^\s<>'\"`\\]*)?)")
TRAILING_URL_CHARS = ".,，。;；:：!?！？)]）】}、"
SYSTEM_URL_HOSTS = {
    "schemas.openxmlformats.org",
    "purl.org",
    "www.w3.org",
    "schemas.microsoft.com",
    "schemas.google.com",
}
SYSTEM_URL_NEEDLES = (
    "openxmlformats.org",
    "schemas.openxmlformats",
    "schemas.microsoft",
    "purl.org",
    "w3.org",
    "wps.cn",
    "kingsoft",
)

SECRET_PATTERNS = [
    ("OpenAI API key", re.compile(r"\bsk-[A-Za-z0-9_-]{32,}\b")),
    ("GitHub token", re.compile(r"\bgh[pousr]_[A-Za-z0-9_]{30,}\b")),
    ("AWS access key", re.compile(r"\bAKIA[0-9A-Z]{16}\b")),
    ("AWS secret key assignment", re.compile(r"(?i)aws(.{0,20})?(secret|private).{0,20}['\"][A-Za-z0-9/+=]{35,}['\"]")),
    ("Private key block", re.compile(r"-----BEGIN (?:RSA |EC |OPENSSH |DSA |)?PRIVATE KEY-----")),
]

TAG_RULES = [
    ("TikTok", ["tiktok", "tk", "抖音"]),
    ("Facebook", ["facebook", "fbgroup", "fb group", "meta ads"]),
    ("YouTube", ["youtube", "youtu.be"]),
    ("Reddit", ["reddit"]),
    ("Twitter/X", ["twitter.com", "x.com", "推特"]),
    ("广告投放", ["广告", "投流", "ads library", "google ads", "adstransparency"]),
    ("七麦", ["七麦", "qimai.cn", "www.qimai.cn"]),
    ("ASO/SEO", ["aso", "seo", "关键词", "抢词", "七麦", "qimai", "site:"]),
    ("RevenueCat", ["revenuecat", "verified.revenuecat"]),
    ("SensorTower", ["sensortower", "sensor tower"]),
    ("社群/分销", ["社群", "知识星球", "zsxq", "分销", "私聊", "二维码"]),
    ("增长打法", ["增长", "裂变", "截流", "借势", "借火", "矩阵", "等风"]),
    ("AI视频/图片", ["ai视频", "ai video", "photo", "image", "图片", "视频", "template"]),
    ("App商店", ["app store", "play.google", "应用市场", "apple", "ios", "android"]),
    ("KOL/达人", ["kol", "达人", "大使", "共创"]),
    ("电商", ["电商", "amazon", "etsy", "shopify", "listing"]),
    ("商业化", ["订阅", "iap", "广告变现", "分销", "联盟", "freemium", "coupon", "付费"]),
    ("竞品监控", ["sitemap", "监控", "追踪", "ads library", "adstransparency", "similarweb"]),
    ("产品思考", ["思考", "复盘", "总结", "想法", "产品"]),
    ("工具/插件", ["插件", "chrome", "github", "replicate", "huggingface", "tool"]),
]

RISK_RULES = [
    ("高风险增长", ["截流", "诱导", "刷榜", "破解", "暴力破解", "薅", "黑产", "灰产", "抄袭"]),
    ("投放/ROI风险", ["投流", "roi", "广告", "投放", "素材"]),
    ("平台规则风险", ["商标", "李鬼", "老二", "蹭词", "抢词"]),
    ("私域/分销信息", ["知识星球", "二维码", "私聊", "微信", "手机号", "姓名", "转账"]),
    ("联盟/追踪链接", ["affiliate", "aff=", "ref=", "invite=", "utm_", "gclid", "gbraid", "campaign"]),
    ("成人/社交敏感", ["dating", "adult", "聊骚", "约约", "骚货"]),
    ("恶搞/拟真敏感", ["prank", "fake chat", "假短信", "fake whatsapp"]),
]

DIMENSION_FALLBACK = "原文未明确，需要结合上下文复盘。"
FIVE_DIMENSIONS = [
    ("dao", "道", "底层逻辑", ["想法", "赚钱方式", "变现方式", "商业化", "流量", "需求", "分类", "特点", "结果", "收益"]),
    ("fa", "法", "方法路径", ["方法", "方式", "路径", "引流", "导入", "SEO", "ASO", "投放", "裂变", "监控", "教程"]),
    ("shu", "术", "具体动作", ["步骤", "操作", "功能", "地址", "网址", "搜索", "复制", "做成", "部署", "上传", "1", "2"]),
    ("qi", "器", "工具平台", ["工具", "产品", "平台", "地址", "网址", "github", "qimai", "sensortower", "app store", "facebook", "tiktok"]),
    ("shi", "势", "趋势窗口", ["趋势", "增长", "流量", "上线", "一个月", "热点", "榜", "trending", "市场", "机会", "窗口"]),
]

DIMENSION_TAG_HINTS = {
    "TikTok": {
        "fa": "方法线索：围绕 TikTok 内容、账号、模板或爆款样本做选题、引流或截流复盘。",
        "shi": "势能线索：短视频流量池和热点素材是这份资料的重要观察对象。",
    },
    "Facebook": {
        "fa": "方法线索：通过 Facebook 社群、广告素材或公开广告库观察可迁移流量。",
        "qi": "器物线索：Facebook/Meta 相关页面可作为素材和投放观察入口。",
    },
    "ASO/SEO": {
        "fa": "方法线索：从关键词、榜单、搜索结果或竞品页面反推需求。",
        "shi": "势能线索：搜索和应用商店排名暴露了长期需求窗口。",
    },
    "七麦": {
        "qi": "器物线索：七麦数据可用于应用榜单、关键词和竞品验证。",
    },
    "RevenueCat": {
        "qi": "器物线索：RevenueCat 相关公开页面可用于订阅与商业化线索复盘。",
        "dao": "底层逻辑：订阅收入和付费验证是判断产品价值的重要证据。",
    },
    "广告投放": {
        "fa": "方法线索：用广告素材、透明库和投放词观察竞品获客方式。",
        "shi": "势能线索：广告持续投放通常意味着市场需求或 ROI 仍值得复盘。",
    },
    "社群/分销": {
        "dao": "底层逻辑：资料、社群入口和分销机制共同承接信任与转化。",
        "fa": "方法线索：用内容或资料做入口，再引导到社群、私域或分销承接。",
    },
    "工具/插件": {
        "qi": "器物线索：插件、代码库、AI 工具和平台链接是复盘具体实现的入口。",
    },
    "竞品监控": {
        "fa": "方法线索：通过 sitemap、广告库、榜单或公开页面持续监控竞品变化。",
    },
    "商业化": {
        "dao": "底层逻辑：原文关注订阅、付费、广告、分销或服务化变现。",
    },
}


def posix(path: Path) -> str:
    return path.as_posix()


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def should_skip(path: Path) -> bool:
    return ".git" in path.parts or path.name == ".DS_Store" or path.name.startswith("._")


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def extract_urls(text: str) -> list[str]:
    urls: list[str] = []
    for match in URL_RE.finditer(text):
        url = match.group(0).rstrip(TRAILING_URL_CHARS)
        if url and not is_system_url(url):
            urls.append(url)
    for match in WWW_RE.finditer(text):
        url = "https://" + match.group(1).rstrip(TRAILING_URL_CHARS)
        if url and not is_system_url(url):
            urls.append(url)
    return urls


def is_system_url(url: str) -> bool:
    normalized = html.unescape(url).lower()
    parsed = urlparse(normalized)
    host = parsed.netloc.lower()
    if host in SYSTEM_URL_HOSTS:
        return True
    return any(needle in normalized for needle in SYSTEM_URL_NEEDLES)


def normalize_url(url: str) -> str:
    cleaned = html.unescape(url).strip().rstrip(TRAILING_URL_CHARS)
    if cleaned.startswith("www."):
        cleaned = "https://" + cleaned
    return cleaned


def unwrap_redirect(url: str) -> str | None:
    parsed = urlparse(url)
    if not parsed.query:
        return None
    query = parse_qs(parsed.query)
    for key in ("q", "url", "u", "target", "redirect", "redirect_url", "destination"):
        values = query.get(key)
        if not values:
            continue
        value = unquote(values[0])
        if value.startswith(("http://", "https://")) and not is_system_url(value):
            return normalize_url(value)
    return None


def classify_domain(domain: str, url: str) -> str:
    lowered = f"{domain} {url}".lower()
    if any(host in lowered for host in ("qimai.cn", "sensortower.com", "apps.apple.com", "play.google.com", "chromewebstore.google.com", "microsoftedge.microsoft.com")):
        return "app_market_intel"
    if any(host in lowered for host in ("tiktok.com", "youtube.com", "facebook.com", "instagram.com", "x.com", "twitter.com", "reddit.com", "bilibili.com")):
        return "social_video"
    if any(host in lowered for host in ("zsxq.com", "t.zsxq.com", "mp.weixin.qq.com", "nas.io")):
        return "community_entry"
    if any(token in lowered for token in ("ads/library", "adstransparency.google.com", "trends.google.com")):
        return "ads_competitor_research"
    if any(host in lowered for host in ("github.com", "huggingface.co", "replicate.com", "openai.com", "workspace.google.com")):
        return "code_ai_infra"
    if any(token in lowered for token in ("affiliate", "aff=", "ref=", "invite=", "utm_", "gclid", "gbraid")):
        return "affiliate_tracking_redirect"
    return "ai_saas_tools"


def risk_tags_for_url(url: str) -> list[str]:
    lowered = url.lower()
    tags = []
    if any(token in lowered for token in ("affiliate", "aff=", "ref=", "invite=")):
        tags.append("联盟/追踪链接")
    if any(token in lowered for token in ("utm_", "gclid", "gbraid", "campaign", "channel")):
        tags.append("联盟/追踪链接")
    if "ads/library" in lowered or "adstransparency" in lowered:
        tags.append("投放/ROI风险")
    return tags


def first_heading(text: str) -> str | None:
    for line in text.splitlines():
        match = re.match(r"^\s*#{1,6}\s+(.+?)\s*$", line)
        if match:
            return match.group(1).strip()
    return None


def excerpt(text: str, limit: int = 220) -> str:
    normalized = re.sub(r"\s+", " ", text).strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[: limit - 1] + "…"


def compact_evidence_line(line: str, limit: int = 92) -> str:
    cleaned = html.unescape(line)
    cleaned = re.sub(r"<[^>]+>", " ", cleaned)
    cleaned = re.sub(r"!\[[^\]]*\]\([^)]+\)", " ", cleaned)
    cleaned = re.sub(r"\[[^\]]+\]\([^)]+\)", " ", cleaned)
    cleaned = re.sub(r"https?://[^\s]+", "链接", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -*#\t")
    if len(cleaned) <= limit:
        return cleaned
    return cleaned[: limit - 1] + "…"


def evidence_lines(text: str, needles: list[str], limit: int = 3) -> list[str]:
    if not text:
        return []
    results: list[str] = []
    seen: set[str] = set()
    lowered_needles = [needle.lower() for needle in needles if needle.strip()]
    for raw_line in text.splitlines():
        line = compact_evidence_line(raw_line)
        if len(line) < 8:
            continue
        lowered = line.lower()
        if not any(needle in lowered for needle in lowered_needles):
            continue
        if line in seen:
            continue
        seen.add(line)
        results.append(line)
        if len(results) >= limit:
            break
    return results


def domain_summary(urls: list[str], limit: int = 5) -> list[str]:
    domains: Counter[str] = Counter()
    for raw_url in urls:
        url = normalize_url(raw_url)
        final_url = unwrap_redirect(url)
        parsed = urlparse(final_url or url)
        domain = parsed.netloc.lower()
        if domain:
            domains[domain] += 1
    return [domain for domain, _count in domains.most_common(limit)]


def dimension_confidence(items: list[str]) -> str:
    return "needs_review" if items == [DIMENSION_FALLBACK] else "evidence"


def build_five_dimensions(
    rel_path: str,
    title: str,
    text: str,
    tags: list[str],
    risk_tags: list[str],
    urls: list[str],
    year: str,
    ext: str,
    attachment: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    tag_hints: dict[str, list[str]] = defaultdict(list)
    for tag in tags:
        for key, hint in DIMENSION_TAG_HINTS.get(tag, {}).items():
            tag_hints[key].append(hint)

    domains = domain_summary(urls)
    dimensions: list[dict[str, Any]] = []
    useful_tags = [tag for tag in tags if tag not in {"未分类", "附件"}]

    for key, label, subtitle, needles in FIVE_DIMENSIONS:
        items: list[str] = []
        items.extend(tag_hints.get(key, [])[:2])

        if key == "dao" and useful_tags:
            items.append(f"主题判断：这份资料主要落在 {'、'.join(useful_tags[:5])}。")
        if key == "fa" and "增长打法" in tags:
            items.append("方法线索：原文可放入“发现机会 → 验证需求 → 承接流量 → 转化复盘”的链路。")
        if key == "shu":
            for line in evidence_lines(text, needles, 3):
                items.append(f"动作证据：{line}")
        elif key not in {"qi"}:
            for line in evidence_lines(text, needles, 2):
                items.append(f"原文证据：{line}")

        if key == "qi":
            if domains:
                items.append(f"关联工具/平台：{'、'.join(domains)}。")
            if attachment:
                items.append(f"附件入口：{attachment.get('kind', ext.lstrip('.') or 'file')}，需要打开原始文件查看细节。")
            for line in evidence_lines(text, needles, 2):
                items.append(f"原文证据：{line}")

        if key == "shi":
            if year != "未标年":
                items.append(f"时间线索：资料归入 {year}，适合和当年的平台热度、榜单和投放环境一起看。")
            if risk_tags:
                items.append(f"边界提醒：原文触及 {'、'.join(risk_tags[:4])}，复盘时先区分学习观察和实际执行。")

        deduped: list[str] = []
        seen: set[str] = set()
        for item in items:
            cleaned = compact_evidence_line(item, 120)
            if not cleaned or cleaned in seen:
                continue
            seen.add(cleaned)
            deduped.append(cleaned)
            if len(deduped) >= 4:
                break
        if not deduped:
            deduped = [DIMENSION_FALLBACK]

        dimensions.append(
            {
                "key": key,
                "label": label,
                "subtitle": subtitle,
                "items": deduped,
                "confidence": dimension_confidence(deduped),
                "source": "generated-from-local-content",
            }
        )

    return dimensions


def detect_tags(rel_path: str, text: str, ext: str) -> list[str]:
    haystack = f"{rel_path}\n{text}".lower()
    tags = []
    for tag, needles in TAG_RULES:
        if any(needle.lower() in haystack for needle in needles):
            tags.append(tag)
    if ext in {".pptx", ".xlsx", ".jpg", ".jpeg", ".png"}:
        tags.append("附件")
    if not tags:
        tags.append("未分类")
    return sorted(set(tags))


def detect_risk_tags(rel_path: str, text: str) -> list[str]:
    haystack = f"{rel_path}\n{text}".lower()
    tags = []
    for tag, needles in RISK_RULES:
        if any(needle.lower() in haystack for needle in needles):
            tags.append(tag)
    return sorted(set(tags))


def infer_year(rel_path: str, text: str) -> str:
    match = re.search(r"(20\d{2})", rel_path)
    if match:
        return match.group(1)
    match = re.search(r"(20\d{2})", text[:500])
    if match:
        return match.group(1)
    return "未标年"


def public_path(rel_path: str) -> str:
    return "raw/" + quote(rel_path, safe="/")


def scan_for_blocking_secrets(source: Path) -> list[dict[str, str]]:
    findings: list[dict[str, str]] = []
    for path in sorted(source.rglob("*")):
        if should_skip(path) or not path.is_file():
            continue
        if path.suffix.lower() not in {".md", ".txt", ".json", ".js", ".html", ".csv", ".xml", ".yml", ".yaml", ".env"}:
            continue
        text = read_text(path)
        for label, pattern in SECRET_PATTERNS:
            if pattern.search(text):
                findings.append({"path": posix(path.relative_to(source)), "type": label})
    return findings


def copy_raw_files(source: Path) -> list[Path]:
    if RAW_DIR.exists():
        shutil.rmtree(RAW_DIR)
    RAW_DIR.mkdir(parents=True)

    copied: list[Path] = []
    for path in sorted(source.rglob("*")):
        if should_skip(path) or not path.is_file():
            continue
        rel = path.relative_to(source)
        destination = RAW_DIR / rel
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(path, destination)
        copied.append(destination)
    return copied


def inspect_xlsx(path: Path) -> dict[str, Any]:
    meta: dict[str, Any] = {"kind": "xlsx", "sheets": []}
    try:
        import openpyxl  # type: ignore

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = []
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row or 0, 5), values_only=True):
                rows.append([str(value) if value is not None else "" for value in row[:8]])
            meta["sheets"].append(
                {
                    "name": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "sample": rows,
                }
            )
    except Exception as error:  # pragma: no cover - diagnostic metadata only.
        meta["error"] = repr(error)
    return meta


def inspect_pptx(path: Path) -> dict[str, Any]:
    meta: dict[str, Any] = {"kind": "pptx", "slides": [], "mediaCount": 0}
    try:
        with zipfile.ZipFile(path) as archive:
            slide_names = sorted(
                [name for name in archive.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
                key=lambda name: int(re.search(r"slide(\d+)\.xml", name).group(1)),  # type: ignore[union-attr]
            )
            meta["mediaCount"] = len([name for name in archive.namelist() if name.startswith("ppt/media/")])
            for slide_name in slide_names:
                xml = archive.read(slide_name)
                tree = ET.fromstring(xml)
                texts = [node.text.strip() for node in tree.iter() if node.tag.endswith("}t") and node.text and node.text.strip()]
                slide_number = int(re.search(r"slide(\d+)\.xml", slide_name).group(1))  # type: ignore[union-attr]
                meta["slides"].append({"number": slide_number, "text": excerpt(" ".join(texts), 260)})
    except Exception as error:  # pragma: no cover - diagnostic metadata only.
        meta["error"] = repr(error)
    return meta


def inspect_attachment(path: Path) -> dict[str, Any] | None:
    ext = path.suffix.lower()
    if ext == ".xlsx":
        return inspect_xlsx(path)
    if ext == ".pptx":
        return inspect_pptx(path)
    if ext in {".jpg", ".jpeg", ".png", ".gif", ".webp"}:
        return {"kind": "image"}
    return None


def extract_zip_urls(path: Path) -> list[str]:
    urls: list[str] = []
    try:
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                if not name.endswith((".xml", ".rels")):
                    continue
                try:
                    text = archive.read(name).decode("utf-8", errors="replace")
                except UnicodeDecodeError:
                    continue
                urls.extend(extract_urls(text))
    except zipfile.BadZipFile:
        return []
    return urls


def extract_attachment_urls(path: Path) -> list[str]:
    ext = path.suffix.lower()
    if ext not in {".xlsx", ".pptx"}:
        return []
    return extract_zip_urls(path)


def build_inventory(source: Path, copied_files: list[Path]) -> dict[str, Any]:
    files: list[dict[str, Any]] = []
    links: list[dict[str, Any]] = []
    unique_urls: dict[str, dict[str, Any]] = {}
    domain_counts: Counter[str] = Counter()
    year_counts: Counter[str] = Counter()
    tag_counts: Counter[str] = Counter()
    top_dir_counts: Counter[str] = Counter()
    attachment_files: list[dict[str, Any]] = []

    for copied in copied_files:
        rel_path = posix(copied.relative_to(RAW_DIR))
        ext = copied.suffix.lower()
        text = ""
        urls: list[str] = []
        headings: list[str] = []
        line_count = None
        if ext in {".md", ".txt"}:
            text = read_text(copied)
            urls = extract_urls(text)
            headings = [match.group(1).strip() for match in re.finditer(r"(?m)^#{1,6}\s+(.+?)\s*$", text)]
            line_count = text.count("\n") + (1 if text else 0)
        elif ext in {".xlsx", ".pptx"}:
            urls = extract_attachment_urls(copied)
        title = first_heading(text) or copied.stem
        year = infer_year(rel_path, text)
        tags = detect_tags(rel_path, text, ext)
        risk_tags = detect_risk_tags(rel_path, text)
        attachment = inspect_attachment(copied)
        five_dimensions = build_five_dimensions(rel_path, title, text, tags, risk_tags, urls, year, ext, attachment)

        for tag in tags:
            tag_counts[tag] += 1
        year_counts[year] += 1
        top_dir_counts[rel_path.split("/")[0]] += 1

        file_item: dict[str, Any] = {
            "path": rel_path,
            "title": title,
            "extension": ext or "",
            "type": "markdown" if ext == ".md" else "attachment" if attachment else "file",
            "size": copied.stat().st_size,
            "sha256": sha256_file(copied),
            "lineCount": line_count,
            "year": year,
            "tags": tags,
            "riskTags": risk_tags,
            "urlCount": len(urls),
            "headings": headings[:20],
            "excerpt": excerpt(text) if text else "",
            "content": text if ext == ".md" else "",
            "publicPath": public_path(rel_path),
            "fiveDimensions": five_dimensions,
        }
        if attachment:
            file_item["attachment"] = attachment
            attachment_files.append(file_item)
        files.append(file_item)

        for raw_url in urls:
            url = normalize_url(raw_url)
            if is_system_url(url):
                continue
            final_url = unwrap_redirect(url)
            parsed = urlparse(final_url or url)
            domain = parsed.netloc.lower()
            if not domain:
                continue
            domain_counts[domain] += 1
            source_risk_tags = sorted(set(risk_tags + risk_tags_for_url(url) + (risk_tags_for_url(final_url) if final_url else [])))
            link_item = {
                "url": url,
                "finalUrl": final_url,
                "domain": domain,
                "group": classify_domain(domain, final_url or url),
                "sourcePath": rel_path,
                "sourceTitle": title,
                "sourcePublicPath": public_path(rel_path),
                "tags": tags,
                "riskTags": source_risk_tags,
                "year": year,
            }
            links.append(link_item)
            canonical = final_url or url
            unique = unique_urls.setdefault(
                canonical,
                {
                    "url": canonical,
                    "rawUrls": set(),
                    "domain": domain,
                    "group": classify_domain(domain, canonical),
                    "count": 0,
                    "sources": [],
                    "tags": set(),
                    "riskTags": set(),
                },
            )
            unique["rawUrls"].add(url)
            unique["count"] += 1
            unique["sources"].append({"path": rel_path, "title": title, "publicPath": public_path(rel_path)})
            unique["tags"].update(tags)
            unique["riskTags"].update(source_risk_tags)

    unique_links = []
    for item in unique_urls.values():
        unique_links.append(
            {
                "url": item["url"],
                "rawUrls": sorted(item["rawUrls"])[:10],
                "domain": item["domain"],
                "group": item["group"],
                "count": item["count"],
                "sources": item["sources"][:20],
                "tags": sorted(item["tags"]),
                "riskTags": sorted(item["riskTags"]),
            }
        )
    unique_links.sort(key=lambda item: (-item["count"], item["domain"], item["url"]))

    markdown_files = [item for item in files if item["extension"] == ".md"]
    stats = {
        "generatedAt": now_iso(),
        "sourcePath": str(source),
        "fileCount": len(files),
        "markdownCount": len(markdown_files),
        "markdownTotalLines": sum(item["lineCount"] or 0 for item in markdown_files),
        "attachmentCount": len(attachment_files),
        "linkOccurrenceCount": len(links),
        "uniqueLinkCount": len(unique_links),
        "yearCounts": dict(sorted(year_counts.items())),
        "tagCounts": dict(sorted(tag_counts.items(), key=lambda pair: (-pair[1], pair[0]))),
        "topDirCounts": dict(sorted(top_dir_counts.items(), key=lambda pair: (-pair[1], pair[0]))),
        "topDomains": [{"domain": domain, "count": count, "group": classify_domain(domain, domain)} for domain, count in domain_counts.most_common(40)],
    }

    return {
        "stats": stats,
        "files": files,
        "links": links,
        "uniqueLinks": unique_links,
        "attachments": attachment_files,
    }


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def write_data_js(inventory: dict[str, Any]) -> None:
    SITE_DIR.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(inventory, ensure_ascii=False, separators=(",", ":"))
    (SITE_DIR / "moneyhunter-data.js").write_text(
        "\"use strict\";\nwindow.moneyHunterContent = " + payload + ";\n",
        encoding="utf-8",
    )


def card_list(items: list[str]) -> str:
    return "\n".join(f"- {item}" for item in items)


def generate_docs(inventory: dict[str, Any]) -> None:
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    stats = inventory["stats"]
    top_domains = "\n".join(f"- `{item['domain']}`：{item['count']} 次" for item in stats["topDomains"][:20])
    top_tags = "\n".join(f"- {tag}：{count} 个文件" for tag, count in list(stats["tagCounts"].items())[:20])
    top_dirs = "\n".join(f"- `{name}`：{count} 个文件" for name, count in stats["topDirCounts"].items())

    overview = f"""# MoneyHunter 全量公开学习库总览

本仓库把 MoneyHunter 当前本地工作树整理成一个公开、可搜索、可下载、可复盘的静态学习库。

## 数据范围

- 源目录：`{stats['sourcePath']}`
- 生成时间：`{stats['generatedAt']}`
- 公开文件数：{stats['fileCount']}
- Markdown 文件数：{stats['markdownCount']}
- Markdown 总行数：{stats['markdownTotalLines']}
- 附件数：{stats['attachmentCount']}
- URL 出现次数：{stats['linkOccurrenceCount']}
- 唯一 URL 数：{stats['uniqueLinkCount']}

## 原始资料入口

- 全量原始资料在 [`raw/`](../raw/)。
- 交互学习首页在 [`site/index.html`](../site/index.html)。
- 公开链接库在 [`site/links.html`](../site/links.html)。

## 顶层目录分布

{top_dirs}

## 主题分布

{top_tags}

## 高频外链域名

{top_domains}

## MoneyHunter 核心框架

这批资料反复围绕一件事：人在固定流量池里出现，被诱饵吸引，跳转到另一个承载物，再通过订阅、广告、资料、社群或分销变现。学习时不要只看“某个术”，要把它放回完整链路里：

`发现热点/竞品 → 验证需求/关键词 → 做承载页面或产品 → 引流/投放 → 转化/付费 → 放大/复盘`

## 推荐学习顺序

1. `总览入门`：社群介绍 → 2025 总结 → 增长打法地图 → 文件索引。
2. `TikTok 爆款路径`：Prank/Meme → TikTok 选词/引流/截流 → 相关案例。
3. `SEO/ASO 反穿路径`：ASO-SEO → 广告热词战法 → 免费之术 → 关键词机会清单。
4. `广告投流路径`：借风之术 → Facebook 传导术 → 无限投放/卡 ROI → 免费导 App。
5. `AI 视频/图片产品路径`：AI 套壳 2.0 → APP 思考 → Prank/Meme → cuttemplate/AI 社区案例。
6. `商业化路径`：付费思路 → RevenueCat 观察 → 免费之术 → 分销/社群商业说明。
7. `竞品监控路径`：sitemap 追踪 → AI 追踪 → TikTok 爆款监控 → 广告热词战法。
8. `风险复核路径`：商标之术 → 借势之术下架案例 → 借风之术 → 刷榜战术 → 风险边界说明。
"""
    (DOCS_DIR / "moneyhunter-overview.md").write_text(overview, encoding="utf-8")

    growth = """# MoneyHunter 增长打法地图

这份资料库的价值不在于某一个孤立链接，而在于它持续记录了“产品、流量池、素材、关键词、投放、社群”之间的迁移关系。

## 核心增长线索

- TikTok / 短视频：观察热点素材、达人账号、模板类产品、邀请码和裂变路径。
- ASO / SEO：从应用排名、关键词、广告词、竞品页面和搜索结果反向找机会。
- 广告投放：通过 Facebook Ads Library、Google Ads Transparency 等公开页面观察素材变化。
- RevenueCat / 收入线索：从公开订阅验证页面和第三方数据源观察收入、增长和产品成熟度。
- 社群 / 分销：把内容、项目拆解、资料下载和私域入口组合成付费组织。
- 产品变化：从 AI 视频、图片、模板、插件、Web-to-App 等方向观察小团队切入点。

## 标签化地图

- 增长阶段：机会发现、关键词/趋势验证、MVP/上站、引流、转化、商业化、规模化、复盘。
- 流量池：TikTok、Google SEO、ASO、Facebook Ads、YouTube、Reddit、Twitter/X、Instagram、Apple/Google Play、插件/应用市场、电商平台、社群/PR。
- 增长手法：借势、借词、截流、免费/no signup、投流、KOL、矩阵、模板化、sitemap 追踪、竞品广告监控、刷榜/限免、裂变/coupon。
- 产品形态：AI 视频、AI 图片、Prank/Meme、AI 助手、安装服务、健康饮食、教育工具、社交/约会、模板资料、插件/Web 工具、AI 社区。
- 商业化：订阅、IAP、广告/激励视频、分销/联盟、付费服务、资料售卖、Freemium、折扣转化。
- 证据来源：七麦、FB Ad Library、TikTok、Sensor Tower、RevenueCat、App Store、官网、社媒帖子、附件。

## 复盘方法

1. 找一个具体文件，先抽出它提到的产品、链接、平台和指标。
2. 判断它是在观察“需求”“素材”“渠道”“收入”还是“组织方式”。
3. 把外链打开核对，记录哪些链接仍可访问，哪些已经失效。
4. 将打法拆成：入口、诱饵、转化、付费、复购或裂变。
5. 标记平台规则和伦理风险，只做学习复盘，不把高风险动作当默认建议。

## 常用检索词

- `tiktok`
- `Facebook`
- `七麦`
- `ASO`
- `RevenueCat`
- `广告`
- `截流`
- `裂变`
- `知识星球`
"""
    (DOCS_DIR / "moneyhunter-growth-map.md").write_text(growth, encoding="utf-8")

    file_lines = ["# MoneyHunter 全量文件索引", "", "以下文件均已公开到 `raw/`，可在交互站中搜索和打开。", ""]
    for item in inventory["files"]:
        tags = "、".join(item["tags"])
        risk = "；风险：" + "、".join(item["riskTags"]) if item["riskTags"] else ""
        file_lines.append(f"- [`{item['path']}`](../{item['publicPath']})：{item['type']}，{item['size']} bytes，标签：{tags}{risk}")
    (DOCS_DIR / "moneyhunter-file-index.md").write_text("\n".join(file_lines) + "\n", encoding="utf-8")

    risk = """# MoneyHunter 风险边界说明

本仓库按用户要求公开 MoneyHunter 的普通资料、附件和外链，并尽量保留原貌。页面中的风险标签只用于学习提示，不代表删除、审查或价值判断。

## 公开原则

- 普通 Markdown、Excel、PPTX、JPG、二维码/图片、社群入口、外部链接全部公开。
- `.git/` 和 `.DS_Store` 不公开，因为它们不是学习资料。
- 明确凭据型密钥不公开，例如 API key、私钥、访问令牌、云服务密钥。

## 风险标签含义

- 高风险增长：包含截流、刷榜、破解、抄袭、诱导等词汇或类似语义。
- 投放/ROI风险：涉及广告素材、投放、ROI、买量测试。
- 平台规则风险：涉及商标、蹭词、抢词、相似产品等容易触碰平台规则的内容。
- 私域/分销信息：涉及社群、二维码、微信、知识星球、姓名/手机号字段、转账等信息。

## 学习建议

1. 把这些内容当作市场观察和增长复盘材料。
2. 不要把高风险动作直接照搬到线上产品或账号。
3. 对涉及平台规则、隐私、广告、支付和知识产权的动作，先查平台政策。
4. 对已经失效或变化的链接，用 Wayback、官方页面或当前平台搜索复核。
"""
    (DOCS_DIR / "moneyhunter-risk-notes.md").write_text(risk, encoding="utf-8")


def generate_readme(inventory: dict[str, Any]) -> None:
    stats = inventory["stats"]
    readme = f"""# MoneyHunter Learning Hub

MoneyHunter 全量公开交互学习库。这个仓库把 MoneyHunter 本地资料夹整理成 GitHub Pages 可访问的静态站，保留原始资料、附件和外链，同时提供搜索、筛选、目录树、链接库和学习拆解文档。

## 在线入口

- 交互学习首页：<https://siuserxiaowei.github.io/moneyhunter-learning-hub/>
- 站点页面：<https://siuserxiaowei.github.io/moneyhunter-learning-hub/site/index.html>
- 公开链接库：<https://siuserxiaowei.github.io/moneyhunter-learning-hub/site/links.html>

## 当前数据

- 公开文件数：{stats['fileCount']}
- Markdown 文件数：{stats['markdownCount']}
- Markdown 总行数：{stats['markdownTotalLines']}
- 附件数：{stats['attachmentCount']}
- URL 出现次数：{stats['linkOccurrenceCount']}
- 唯一 URL 数：{stats['uniqueLinkCount']}
- 生成时间：`{stats['generatedAt']}`

## 目录

- [`raw/`](raw/)：MoneyHunter 原始资料镜像，保留目录结构。
- [`raw/index.html`](raw/index.html)：GitHub Pages 可打开的原始资料目录页。
- [`site/`](site/)：交互静态站。
- [`docs/`](docs/)：学习化拆解文档。
- [`data/`](data/)：manifest 和链接索引 JSON。
- [`scripts/`](scripts/)：导入与校验脚本。

## 复跑导入

```bash
python3 scripts/import_moneyhunter.py
python3 scripts/verify_package.py
```

## 公开边界

本仓库用于学习、复盘和资料查阅。普通资料、附件、二维码/图片、社群入口和外链按要求公开；`.git/`、`.DS_Store` 和明确凭据型密钥不公开。
"""
    (ROOT / "README.md").write_text(readme, encoding="utf-8")


def generate_raw_index(inventory: dict[str, Any]) -> None:
    page = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>MoneyHunter Raw Archive</title>
  <meta name="description" content="MoneyHunter 全量原始资料目录，支持路径、类型和目录筛选。">
  <link rel="icon" href="data:,">
  <link rel="stylesheet" href="../site/styles.css">
</head>
<body data-page="raw" data-root-prefix="../">
  <header class="site-header">
    <nav class="nav" aria-label="主导航">
      <a class="brand" href="../index.html"><span class="brand-mark">MH</span><span>MoneyHunter Raw Archive</span></a>
      <div class="nav-links">
        <a href="../index.html">学习库</a>
        <a href="../site/links.html">链接库</a>
        <a class="active" href="index.html">Raw</a>
        <a href="https://github.com/siuserxiaowei/moneyhunter-learning-hub/tree/main/raw" target="_blank" rel="noreferrer">GitHub Raw</a>
      </div>
    </nav>
    <section class="masthead">
      <div>
        <p class="eyebrow">Original Files</p>
        <h1>MoneyHunter 原始资料目录</h1>
      </div>
      <p class="lede">{len(inventory['files'])} 个公开文件 · 保留原始目录结构和中文文件名</p>
    </section>
  </header>
  <main>
    <section class="raw-toolbar" aria-label="Raw 资料筛选">
      <label class="search-box">
        <span>路径搜索</span>
        <input id="rawSearchInput" type="search" placeholder="输入文件名、目录、标签">
      </label>
      <label>
        <span>文件类型</span>
        <select id="rawTypeFilter"></select>
      </label>
      <span class="result-note" id="rawCount"></span>
    </section>
    <section class="raw-grid">
      <aside class="side-rail">
        <div class="section-head">
          <h2>目录</h2>
          <span id="rawTreeCount"></span>
        </div>
        <div class="tree compact-tree" id="rawDirectoryTree"></div>
      </aside>
      <div class="table-panel">
        <div class="section-head">
          <h2>全量文件</h2>
          <a class="text-link" href="https://github.com/siuserxiaowei/moneyhunter-learning-hub/tree/main/raw" target="_blank" rel="noreferrer">GitHub Raw</a>
        </div>
        <div class="responsive-table">
          <table class="raw-table">
            <thead><tr><th>文件</th><th>类型</th><th>年份</th><th>标签</th><th>大小</th></tr></thead>
            <tbody id="rawTableBody"></tbody>
          </table>
        </div>
      </div>
    </section>
  </main>
  <footer class="footer">
    <p>MoneyHunter Learning Hub · Raw archive.</p>
  </footer>
  <script src="../site/moneyhunter-data.js"></script>
  <script src="../site/app.js"></script>
</body>
</html>
"""
    (RAW_DIR / "index.html").write_text(page, encoding="utf-8")


def html_page(title: str, body_class: str, links_page: bool = False, asset_prefix: str = "site/") -> str:
    is_site_page = asset_prefix == ""
    root_prefix = "../" if is_site_page else ""
    brand_href = "../index.html" if is_site_page else "index.html"
    home_href = "index.html"
    links_href = "links.html" if is_site_page else "site/links.html"
    docs_href = f"{root_prefix}docs/moneyhunter-overview.md"
    raw_href = f"{root_prefix}raw/"
    data_src = f"{asset_prefix}moneyhunter-data.js"
    css_href = f"{asset_prefix}styles.css"
    app_src = f"{asset_prefix}app.js"
    page_kind = "links" if links_page else "home"
    active_home = ' class="active"' if not links_page else ""
    active_links = ' class="active"' if links_page else ""
    if links_page:
        main_content = f"""
    <section class="link-summary">
      <div>
        <p class="eyebrow">Public Link Index</p>
        <h1>MoneyHunter 完整链接库</h1>
      </div>
      <p class="lede">公开收录全部 URL，按域名、来源文件和主题标签检索。</p>
    </section>
    <section class="stats-strip" id="stats" aria-label="链接统计"></section>
    <section class="link-toolbar" id="explore" aria-label="链接筛选">
      <label class="search-box">
        <span>搜索链接</span>
        <input id="linkSearchInput" type="search" placeholder="域名、URL、来源文件、标签">
      </label>
      <label>
        <span>域名</span>
        <select id="domainFilter"></select>
      </label>
      <label>
        <span>分组</span>
        <select id="groupFilter"></select>
      </label>
      <span class="result-note" id="linkCount"></span>
    </section>
    <section class="link-domain-band">
      <div class="section-head">
        <h2>高频域名</h2>
        <button class="ghost-button" id="clearLinkFilters" type="button">清空筛选</button>
      </div>
      <div class="chip-row" id="linkDomainChips"></div>
    </section>
    <section class="link-index">
      <div class="section-head">
        <h2>链接结果</h2>
        <div class="pager" aria-label="链接分页">
          <button class="ghost-button" id="linkPrev" type="button">上一页</button>
          <span id="linkPager"></span>
          <button class="ghost-button" id="linkNext" type="button">下一页</button>
        </div>
      </div>
      <div class="link-list" id="linkTableBody"></div>
    </section>
"""
    else:
        main_content = f"""
    <section class="stats-strip" id="stats" aria-label="资料统计"></section>
    <section class="desk-toolbar" id="explore" aria-label="资料筛选">
      <label class="search-box">
        <span>全文搜索</span>
        <input id="searchInput" type="search" placeholder="tiktok、revenuecat、ASO、七麦、Facebook、知识星球">
      </label>
      <label>
        <span>年份</span>
        <select id="yearFilter"></select>
      </label>
      <label>
        <span>主题</span>
        <select id="tagFilter"></select>
      </label>
      <label>
        <span>风险标签</span>
        <select id="riskFilter"></select>
      </label>
    </section>
    <section class="workspace-grid">
      <aside class="side-rail">
        <div class="section-head">
          <h2>目录树</h2>
          <span id="treeCount"></span>
        </div>
        <div class="tree" id="directoryTree"></div>
      </aside>
      <section class="result-pane">
        <div class="section-head">
          <h2>文件检索</h2>
          <span id="resultCount"></span>
        </div>
        <div class="file-list" id="fileResults"></div>
      </section>
      <aside class="detail-pane" id="fileDetail">
        <div class="empty-detail">
          <p class="eyebrow">File Detail</p>
          <h2>选择一份资料</h2>
          <p>右侧会显示原文摘要、标签、外链和公开打开入口。</p>
        </div>
      </aside>
    </section>
    <section class="learning-band">
      <div class="section-head">
        <h2>学习拆解</h2>
        <a class="text-link" href="{docs_href}">打开完整文档</a>
      </div>
      <div class="study-grid">
        <a href="{root_prefix}docs/moneyhunter-overview.md">总览路径</a>
        <a href="{root_prefix}docs/moneyhunter-growth-map.md">增长打法地图</a>
        <a href="{root_prefix}docs/moneyhunter-file-index.md">全量文件索引</a>
        <a href="{root_prefix}docs/moneyhunter-risk-notes.md">风险边界说明</a>
      </div>
    </section>
    <section class="asset-grid">
      <div class="asset-section">
        <div class="section-head">
          <h2>附件库</h2>
          <span id="attachmentCount"></span>
        </div>
        <div class="attachment-list" id="attachmentGrid"></div>
      </div>
      <div class="asset-section">
        <div class="section-head">
          <h2>高频外链域名</h2>
          <a class="text-link" href="{links_href}">完整链接库</a>
        </div>
        <div class="domain-list" id="domainGrid"></div>
      </div>
    </section>
"""
    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{html.escape(title)}</title>
  <meta name="description" content="MoneyHunter 全量公开交互学习库，支持原文、附件、外链、目录和全文搜索。">
  <link rel="icon" href="data:,">
  <link rel="stylesheet" href="{css_href}">
</head>
<body data-page="{page_kind}" data-root-prefix="{root_prefix}">
  <header class="site-header">
    <nav class="nav" aria-label="主导航">
      <a class="brand" href="{brand_href}">
        <span class="brand-mark">MH</span>
        <span>MoneyHunter Learning Hub</span>
      </a>
      <div class="nav-links">
        <a{active_home} href="{home_href}">学习库</a>
        <a{active_links} href="{links_href}">链接库</a>
        <a href="{docs_href}">文档</a>
        <a href="{raw_href}">Raw</a>
        <a href="https://github.com/siuserxiaowei/moneyhunter-learning-hub" target="_blank" rel="noreferrer">GitHub</a>
      </div>
    </nav>
    <section class="masthead">
      <div>
        <p class="eyebrow">Full Public Archive</p>
        <h1>{html.escape(title)}</h1>
      </div>
      <div class="masthead-actions">
        <a class="button primary" href="#explore">开始检索</a>
        <a class="button secondary" href="{raw_href}">Raw</a>
        <a class="button secondary" href="{docs_href}">Docs</a>
      </div>
    </section>
  </header>
  <main>
{main_content}
  </main>
  <footer class="footer">
    <p>MoneyHunter Learning Hub · 全量公开资料库 · GitHub Pages 静态站。</p>
  </footer>
  <script src="{data_src}"></script>
  <script src="{app_src}"></script>
</body>
</html>
"""


def generate_site() -> None:
    SITE_DIR.mkdir(parents=True, exist_ok=True)
    (SITE_DIR / "index.html").write_text(html_page("MoneyHunter 全量公开交互学习库", "", asset_prefix=""), encoding="utf-8")
    (SITE_DIR / "links.html").write_text(html_page("MoneyHunter 公开链接库", "", links_page=True, asset_prefix=""), encoding="utf-8")
    (ROOT / "index.html").write_text(html_page("MoneyHunter 全量公开交互学习库", "", asset_prefix="site/"), encoding="utf-8")

    css = """
:root {
  color-scheme: light;
  --ink: #18211f;
  --muted: #66736d;
  --faint: #8a9690;
  --bg: #f5f7f8;
  --surface: #ffffff;
  --surface-soft: #eef4f2;
  --line: #d7dee2;
  --line-strong: #b9c4ca;
  --green: #0d6b54;
  --blue: #245f9f;
  --red: #a84834;
  --amber: #b98518;
  --header: #14211d;
}

* { box-sizing: border-box; }
html { scroll-behavior: smooth; }
body {
  margin: 0;
  font-family: "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
  color: var(--ink);
  background: var(--bg);
  line-height: 1.55;
}
a { color: inherit; }
button, input, select { font: inherit; }
button { cursor: pointer; }
h1, h2, h3, p { margin: 0; overflow-wrap: anywhere; }
h1, h2, h3 { line-height: 1.18; letter-spacing: 0; }

.site-header {
  color: #fbfffd;
  background: var(--header);
  border-bottom: 4px solid #d5a735;
}
.nav,
.masthead,
main {
  width: min(1320px, calc(100% - 36px));
  margin: 0 auto;
}
.nav {
  min-height: 66px;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 20px;
}
.brand {
  display: inline-flex;
  align-items: center;
  gap: 11px;
  color: #fff;
  font-weight: 800;
  text-decoration: none;
}
.brand-mark {
  width: 36px;
  height: 36px;
  display: inline-grid;
  place-items: center;
  border: 1px solid rgba(255,255,255,.36);
  border-radius: 6px;
  background: rgba(255,255,255,.08);
  font-size: .85rem;
}
.nav-links {
  display: flex;
  flex-wrap: wrap;
  gap: 4px;
  justify-content: flex-end;
}
.nav-links a {
  min-height: 34px;
  display: inline-flex;
  align-items: center;
  padding: 0 10px;
  border-radius: 6px;
  color: rgba(255,255,255,.78);
  text-decoration: none;
  font-size: .94rem;
}
.nav-links a:hover,
.nav-links a.active {
  background: rgba(255,255,255,.13);
  color: #fff;
}
.masthead {
  display: flex;
  align-items: flex-end;
  justify-content: space-between;
  gap: 24px;
  padding: 28px 0 30px;
}
.eyebrow {
  color: #e1bc55;
  font-size: .76rem;
  font-weight: 800;
  letter-spacing: .08em;
  text-transform: uppercase;
}
h1 {
  max-width: 860px;
  margin-top: 8px;
  font-size: clamp(1.9rem, 4.2vw, 3.45rem);
  font-weight: 900;
}
.lede {
  max-width: 560px;
  color: rgba(255,255,255,.78);
  font-size: .98rem;
}
.masthead-actions,
.detail-actions,
.pager {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 8px;
}
.button,
.ghost-button {
  min-height: 36px;
  display: inline-flex;
  align-items: center;
  justify-content: center;
  border-radius: 6px;
  text-decoration: none;
  font-weight: 800;
  border: 1px solid transparent;
  padding: 0 12px;
}
.button.primary { background: #e3b63f; color: #16201c; }
.button.secondary { background: #fff; color: var(--ink); border-color: var(--line); }
.site-header .button.secondary {
  color: #fff;
  border-color: rgba(255,255,255,.32);
  background: rgba(255,255,255,.08);
}
.ghost-button {
  background: #fff;
  color: var(--ink);
  border-color: var(--line);
}
.ghost-button:disabled { opacity: .45; cursor: not-allowed; }

main { padding: 18px 0 54px; }
.stats-strip {
  display: grid;
  grid-template-columns: repeat(6, minmax(0, 1fr));
  gap: 10px;
  margin-bottom: 14px;
}
.stat-card {
  min-height: 76px;
  padding: 13px 14px;
  background: var(--surface);
  border: 1px solid var(--line);
  border-left: 4px solid var(--green);
  border-radius: 6px;
}
.stat-card span {
  display: block;
  color: var(--muted);
  font-size: .82rem;
  font-weight: 700;
}
.stat-card strong {
  display: block;
  margin-top: 4px;
  font-size: 1.55rem;
  line-height: 1;
}

.desk-toolbar,
.link-toolbar,
.raw-toolbar {
  position: sticky;
  top: 0;
  z-index: 5;
  display: grid;
  gap: 10px;
  align-items: end;
  padding: 12px;
  margin-bottom: 14px;
  background: rgba(245,247,248,.96);
  border: 1px solid var(--line);
  border-radius: 6px;
  backdrop-filter: blur(8px);
}
.desk-toolbar { grid-template-columns: minmax(260px, 1.7fr) repeat(3, minmax(130px, .62fr)); }
.link-toolbar { grid-template-columns: minmax(300px, 1.5fr) minmax(170px, .7fr) minmax(170px, .7fr) auto; }
.raw-toolbar { grid-template-columns: minmax(280px, 1.4fr) minmax(160px, .45fr) auto; }
label span {
  display: block;
  margin-bottom: 5px;
  color: var(--muted);
  font-size: .78rem;
  font-weight: 800;
}
input,
select {
  width: 100%;
  min-height: 38px;
  padding: 0 10px;
  color: var(--ink);
  background: #fff;
  border: 1px solid var(--line-strong);
  border-radius: 6px;
}
input:focus,
select:focus {
  outline: 2px solid rgba(36,95,159,.22);
  border-color: var(--blue);
}
.result-note {
  color: var(--muted);
  font-size: .9rem;
  font-weight: 800;
  white-space: nowrap;
}

.workspace-grid {
  display: grid;
  grid-template-columns: minmax(210px, 250px) minmax(300px, 420px) minmax(420px, 1fr);
  gap: 14px;
  align-items: start;
}
.raw-grid {
  display: grid;
  grid-template-columns: minmax(240px, 310px) minmax(0, 1fr);
  gap: 14px;
  align-items: start;
}
.side-rail,
.result-pane,
.detail-pane,
.asset-section,
.learning-band,
.link-domain-band,
.link-index,
.table-panel {
  background: var(--surface);
  border: 1px solid var(--line);
  border-radius: 6px;
}
.side-rail,
.detail-pane {
  position: sticky;
  top: 84px;
  max-height: calc(100vh - 102px);
  overflow: auto;
}
.side-rail,
.result-pane,
.detail-pane,
.asset-section,
.learning-band,
.link-domain-band,
.link-index,
.table-panel {
  padding: 14px;
}
.section-head {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  padding-bottom: 10px;
  margin-bottom: 10px;
  border-bottom: 1px solid var(--line);
}
.section-head h2 {
  font-size: 1.04rem;
  font-weight: 900;
}
.section-head span,
.muted {
  color: var(--muted);
  font-size: .88rem;
}

.tree {
  display: grid;
  gap: 2px;
}
.tree button {
  width: 100%;
  min-height: 32px;
  padding: 6px 8px 6px calc(8px + var(--depth, 0) * 14px);
  color: var(--ink);
  background: transparent;
  border: 0;
  border-radius: 5px;
  text-align: left;
  font-size: .88rem;
}
.tree button:hover,
.tree button.is-active {
  color: var(--green);
  background: var(--surface-soft);
}
.compact-tree button { font-size: .84rem; }

.file-list,
.attachment-list,
.domain-list,
.link-list {
  display: grid;
  gap: 8px;
}
.file-row {
  width: 100%;
  display: grid;
  grid-template-columns: minmax(0, 1fr);
  gap: 8px;
  padding: 12px;
  color: inherit;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 6px;
  text-align: left;
}
.file-row:hover,
.file-row.is-selected {
  border-color: rgba(13,107,84,.58);
  background: #f7fbf9;
}
.file-main {
  min-width: 0;
  display: block;
}
.file-title {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 8px;
  margin-bottom: 5px;
}
.file-title strong {
  font-size: 1rem;
  line-height: 1.3;
}
.file-path,
.link-url {
  display: block;
  color: var(--blue);
  overflow-wrap: anywhere;
  word-break: break-word;
}
.file-excerpt {
  display: block;
  margin-top: 7px;
  color: #394541;
  font-size: .92rem;
  overflow-wrap: anywhere;
  word-break: break-word;
  display: -webkit-box;
  -webkit-line-clamp: 2;
  -webkit-box-orient: vertical;
  overflow: hidden;
}
.row-meta {
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  justify-content: flex-start;
}
.meta-row {
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  margin: 9px 0;
}
.pill {
  min-height: 22px;
  display: inline-flex;
  align-items: center;
  padding: 0 7px;
  color: var(--green);
  background: #e8f3ee;
  border-radius: 999px;
  font-size: .74rem;
  font-weight: 800;
}
.pill.neutral { color: var(--muted); background: #eef1f3; }
.pill.risk { color: var(--red); background: #fff0ea; }
.dimension-mini {
  display: inline-flex;
  align-items: center;
  gap: 2px;
  padding: 2px;
  color: #fff;
  background: #17211e;
  border: 1px solid rgba(255,255,255,.2);
  border-radius: 6px;
  font-size: .72rem;
  font-weight: 900;
  letter-spacing: 0;
}
.dimension-mini span {
  min-width: 19px;
  min-height: 19px;
  display: inline-grid;
  place-items: center;
  border-radius: 4px;
  background: rgba(255,255,255,.12);
}
.dimension-mini .needs-review {
  color: #d7c8a6;
  background: rgba(255,255,255,.06);
}
.empty-detail {
  display: grid;
  gap: 8px;
  min-height: 240px;
  align-content: center;
  color: var(--muted);
}
.detail-pane h2 {
  font-size: 1.22rem;
}
.detail-hero {
  padding: 16px;
  margin: -2px -2px 14px;
  color: #fff;
  background: linear-gradient(135deg, #17211e, #25473d 58%, #5b3d1a);
  border-radius: 6px;
}
.detail-hero .file-path {
  margin-top: 8px;
  color: #cddfd7;
}
.detail-kicker {
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 8px;
  margin-bottom: 10px;
  color: #e9d18b;
  font-size: .78rem;
  font-weight: 900;
}
.detail-hero .pill.neutral {
  color: #dfe8e4;
  background: rgba(255,255,255,.12);
}
.dimension-title {
  display: flex;
  align-items: baseline;
  justify-content: space-between;
  gap: 12px;
  margin: 14px 0 10px;
  padding-bottom: 8px;
  border-bottom: 1px solid var(--line);
}
.dimension-title span {
  color: var(--amber);
  font-size: .76rem;
  font-weight: 900;
  letter-spacing: .08em;
}
.dimension-title strong {
  font-size: 1.06rem;
}
.dimension-grid {
  display: grid;
  grid-template-columns: repeat(2, minmax(0, 1fr));
  gap: 10px;
  margin-bottom: 14px;
}
.dimension-card {
  min-height: 156px;
  padding: 12px;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 6px;
  box-shadow: 0 10px 24px rgba(24, 33, 31, .06);
}
.dimension-card:nth-child(1) { border-top: 4px solid #263a33; }
.dimension-card:nth-child(2) { border-top: 4px solid #0d6b54; }
.dimension-card:nth-child(3) { border-top: 4px solid #b98518; }
.dimension-card:nth-child(4) { border-top: 4px solid #245f9f; }
.dimension-card:nth-child(5) { border-top: 4px solid #8d3d32; }
.dimension-card.needs-review {
  background: #faf8f1;
}
.dimension-head {
  display: grid;
  grid-template-columns: 34px minmax(0, 1fr);
  gap: 9px;
  align-items: start;
  margin-bottom: 10px;
}
.dimension-label {
  width: 34px;
  height: 34px;
  display: inline-grid;
  place-items: center;
  color: #fff;
  background: var(--header);
  border-radius: 6px;
  font-size: 1.04rem;
  font-weight: 900;
}
.dimension-head strong,
.dimension-head em {
  display: block;
}
.dimension-head em {
  margin-top: 2px;
  color: var(--muted);
  font-size: .74rem;
  font-style: normal;
  font-weight: 800;
}
.dimension-card ul {
  display: grid;
  gap: 7px;
  margin: 0;
  padding-left: 18px;
  color: #33423d;
  font-size: .88rem;
}
.source-preview {
  margin-top: 12px;
  border: 1px solid var(--line);
  border-radius: 6px;
  background: #fff;
}
.source-preview summary {
  padding: 11px 12px;
  color: var(--ink);
  font-weight: 900;
  cursor: pointer;
}
.detail-pane pre {
  max-height: 390px;
  overflow: auto;
  white-space: pre-wrap;
  overflow-wrap: anywhere;
  padding: 12px;
  margin: 0;
  background: #f7faf9;
  border-top: 1px solid var(--line);
  border-right: 0;
  border-bottom: 0;
  border-left: 0;
  border-radius: 0 0 6px 6px;
  font-size: .86rem;
  line-height: 1.58;
}
.source-list {
  display: grid;
  gap: 6px;
  padding-left: 18px;
}
.source-list a { overflow-wrap: anywhere; }
.text-link {
  color: var(--blue);
  font-weight: 800;
  text-decoration: none;
}
.text-link:hover { text-decoration: underline; }

.learning-band,
.asset-grid,
.link-domain-band,
.link-index {
  margin-top: 14px;
}
.study-grid {
  display: grid;
  grid-template-columns: repeat(4, minmax(0, 1fr));
  gap: 8px;
}
.study-grid a,
.attachment-card,
.domain-card,
.link-card {
  display: block;
  padding: 12px;
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 6px;
  text-decoration: none;
}
.study-grid a {
  color: var(--green);
  font-weight: 900;
}
.asset-grid {
  display: grid;
  grid-template-columns: minmax(0, 1fr) minmax(300px, .7fr);
  gap: 14px;
}
.attachment-card h3,
.domain-card strong,
.link-card h3 {
  display: block;
  margin-bottom: 5px;
  font-size: .98rem;
  line-height: 1.32;
}
.domain-card {
  display: grid;
  grid-template-columns: minmax(0, 1fr) auto;
  gap: 8px;
  align-items: center;
}
.domain-card span { color: var(--muted); font-weight: 800; }

.link-summary {
  display: flex;
  justify-content: space-between;
  gap: 20px;
  align-items: end;
  padding: 6px 0 16px;
}
.link-summary h1 { color: var(--ink); }
.link-summary .lede { color: var(--muted); }
.chip-row {
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
}
.domain-chip {
  min-height: 32px;
  padding: 0 10px;
  color: var(--ink);
  background: #fff;
  border: 1px solid var(--line);
  border-radius: 999px;
  font-weight: 800;
}
.domain-chip:hover,
.domain-chip.is-active {
  color: var(--green);
  border-color: rgba(13,107,84,.58);
  background: #f2faf6;
}
.link-card {
  display: grid;
  grid-template-columns: minmax(0, 1fr) 160px;
  gap: 14px;
}
.link-card > div {
  min-width: 0;
}
.link-card .link-url {
  display: block;
  margin-bottom: 6px;
  font-weight: 800;
  text-decoration: none;
}
.link-context {
  color: #3d4844;
  font-size: .9rem;
}
.link-side {
  display: grid;
  gap: 7px;
  align-content: start;
  justify-items: start;
  min-width: 0;
  color: var(--muted);
  font-size: .86rem;
}
.link-side > *,
.link-tags {
  max-width: 100%;
  min-width: 0;
  overflow-wrap: anywhere;
}
.link-tags {
  display: flex;
  flex-wrap: wrap;
  gap: 4px;
}

.responsive-table {
  width: 100%;
  overflow-x: auto;
}
.raw-table {
  width: 100%;
  min-width: 760px;
  border-collapse: collapse;
  font-size: .9rem;
}
.raw-table th,
.raw-table td {
  padding: 10px 8px;
  border-bottom: 1px solid var(--line);
  vertical-align: top;
  text-align: left;
}
.raw-table th {
  color: var(--muted);
  font-size: .76rem;
  letter-spacing: .06em;
  text-transform: uppercase;
}
.raw-table a { color: var(--blue); font-weight: 800; text-decoration: none; }
.raw-table a:hover { text-decoration: underline; }
.raw-path-cell { max-width: 420px; overflow-wrap: anywhere; }

.footer {
  padding: 28px 18px;
  color: var(--muted);
  text-align: center;
  border-top: 1px solid var(--line);
}

@media (max-width: 1180px) {
  .stats-strip { grid-template-columns: repeat(3, minmax(0, 1fr)); }
  .workspace-grid { grid-template-columns: minmax(210px, 260px) minmax(0, 1fr); }
  .detail-pane {
    position: static;
    grid-column: 1 / -1;
    max-height: none;
  }
  .asset-grid { grid-template-columns: 1fr; }
}

@media (max-width: 860px) {
  .nav,
  .masthead,
  main {
    width: min(100% - 24px, 1320px);
  }
  .nav,
  .masthead,
  .link-summary {
    align-items: flex-start;
    flex-direction: column;
  }
  .desk-toolbar,
  .link-toolbar,
  .raw-toolbar,
  .workspace-grid,
  .raw-grid {
    grid-template-columns: 1fr;
  }
  .side-rail {
    position: static;
    max-height: 280px;
  }
  .study-grid { grid-template-columns: repeat(2, minmax(0, 1fr)); }
  .link-card { grid-template-columns: 1fr; }
  .row-meta { justify-content: flex-start; }
  .dimension-grid { grid-template-columns: 1fr; }
  .raw-table {
    min-width: 0;
    border-collapse: separate;
    border-spacing: 0 8px;
  }
  .raw-table thead { display: none; }
  .raw-table,
  .raw-table tbody,
  .raw-table tr,
  .raw-table td {
    display: block;
    width: 100%;
  }
  .raw-table tr {
    padding: 8px 10px;
    border: 1px solid var(--line);
    border-radius: 6px;
    background: #fff;
  }
  .raw-table td {
    display: grid;
    grid-template-columns: 76px minmax(0, 1fr);
    gap: 10px;
    padding: 6px 0;
    border-bottom: 0;
  }
  .raw-table td::before {
    content: attr(data-label);
    color: var(--muted);
    font-weight: 800;
  }
}

@media (max-width: 560px) {
  .stats-strip,
  .study-grid {
    grid-template-columns: 1fr;
  }
  .file-row {
    grid-template-columns: 1fr;
  }
  .masthead-actions,
  .detail-actions,
  .pager {
    width: 100%;
  }
  .button,
  .ghost-button {
    flex: 1 1 auto;
  }
}
"""
    (SITE_DIR / "styles.css").write_text(css, encoding="utf-8")

    app = """
"use strict";

const data = window.moneyHunterContent || {};
const stats = data.stats || {};
const files = data.files || [];
const links = data.links || [];
const uniqueLinks = data.uniqueLinks || [];
const attachments = data.attachments || [];
const rootPrefix = document.body.dataset.rootPrefix || "";
const fileByPath = new Map(files.map(file => [file.path, file]));
const numberFormatter = new Intl.NumberFormat("zh-CN");

const searchInput = document.querySelector("#searchInput");
const yearFilter = document.querySelector("#yearFilter");
const tagFilter = document.querySelector("#tagFilter");
const riskFilter = document.querySelector("#riskFilter");
const fileResults = document.querySelector("#fileResults");
const resultCount = document.querySelector("#resultCount");
const directoryTree = document.querySelector("#directoryTree");
const treeCount = document.querySelector("#treeCount");
const fileDetail = document.querySelector("#fileDetail");
const attachmentGrid = document.querySelector("#attachmentGrid");
const attachmentCount = document.querySelector("#attachmentCount");
const domainGrid = document.querySelector("#domainGrid");

const linkSearchInput = document.querySelector("#linkSearchInput");
const domainFilter = document.querySelector("#domainFilter");
const groupFilter = document.querySelector("#groupFilter");
const linkTableBody = document.querySelector("#linkTableBody");
const linkCount = document.querySelector("#linkCount");
const linkDomainChips = document.querySelector("#linkDomainChips");
const clearLinkFilters = document.querySelector("#clearLinkFilters");
const linkPrev = document.querySelector("#linkPrev");
const linkNext = document.querySelector("#linkNext");
const linkPager = document.querySelector("#linkPager");

const rawSearchInput = document.querySelector("#rawSearchInput");
const rawTypeFilter = document.querySelector("#rawTypeFilter");
const rawDirectoryTree = document.querySelector("#rawDirectoryTree");
const rawTreeCount = document.querySelector("#rawTreeCount");
const rawTableBody = document.querySelector("#rawTableBody");
const rawCount = document.querySelector("#rawCount");

const groupLabels = {
  app_market_intel: "应用/数据",
  social_video: "社媒/视频",
  community_entry: "社群入口",
  ads_competitor_research: "广告/竞品",
  code_ai_infra: "代码/AI",
  affiliate_tracking_redirect: "联盟/追踪",
  ai_saas_tools: "AI/SaaS"
};

let activePath = "";
let selectedFilePath = "";
let rawActivePath = "";
let linkPage = 1;
const linkPageSize = 80;

function byId(id) {
  return document.querySelector(id);
}

function escapeHtml(value) {
  return String(value || "")
    .replaceAll("&", "&amp;")
    .replaceAll("<", "&lt;")
    .replaceAll(">", "&gt;")
    .replaceAll('"', "&quot;");
}

function formatNumber(value) {
  return numberFormatter.format(value || 0);
}

function formatBytes(bytes) {
  if (!bytes) return "0 B";
  const units = ["B", "KB", "MB", "GB"];
  let value = bytes;
  let unit = 0;
  while (value >= 1024 && unit < units.length - 1) {
    value /= 1024;
    unit += 1;
  }
  return `${value.toFixed(value >= 10 || unit === 0 ? 0 : 1)} ${units[unit]}`;
}

function fileHref(file) {
  return `${rootPrefix}${file.publicPath}`;
}

function githubHref(file) {
  return `https://github.com/siuserxiaowei/moneyhunter-learning-hub/blob/main/${file.publicPath}`;
}

function statCards() {
  const target = byId("#stats");
  if (!target) return;
  const items = [
    ["公开文件", stats.fileCount],
    ["Markdown", stats.markdownCount],
    ["总行数", stats.markdownTotalLines],
    ["附件", stats.attachmentCount],
    ["URL 出现", stats.linkOccurrenceCount],
    ["唯一链接", stats.uniqueLinkCount]
  ];
  target.innerHTML = items
    .map(([label, value]) => `<article class="stat-card"><span>${label}</span><strong>${formatNumber(value)}</strong></article>`)
    .join("");
}

function fillSelect(select, options, allLabel, labeler = value => value) {
  if (!select) return;
  const optionHtml = options.map(option => {
    const label = labeler(option);
    return `<option value="${escapeHtml(option)}">${escapeHtml(label)}</option>`;
  });
  select.innerHTML = [`<option value="">${allLabel}</option>`].concat(optionHtml).join("");
}

function pillList(items, risk = false) {
  return (items || [])
    .slice(0, 8)
    .map(item => `<span class="pill${risk ? " risk" : ""}">${escapeHtml(item)}</span>`)
    .join("");
}

function truncateText(value, limit = 120) {
  const text = String(value || "").replace(/\s+/g, " ").trim();
  if (text.length <= limit) return text;
  return `${text.slice(0, limit - 1)}…`;
}

function dimensionMini(file) {
  const dimensions = file.fiveDimensions || [];
  return `<span class="dimension-mini" aria-label="道法术器势">${
    dimensions.map(item => `<span class="${item.confidence === "needs_review" ? "needs-review" : ""}">${escapeHtml(item.label)}</span>`).join("")
  }</span>`;
}

function dimensionCards(file) {
  const dimensions = file.fiveDimensions || [];
  if (!dimensions.length) return "";
  return `<section class="dimension-grid" aria-label="道法术器势拆解">${dimensions.map(item => {
    const body = (item.items || []).map(point => `<li>${escapeHtml(point)}</li>`).join("");
    const review = item.confidence === "needs_review" ? " needs-review" : "";
    return `<article class="dimension-card${review}">
      <div class="dimension-head">
        <span class="dimension-label">${escapeHtml(item.label)}</span>
        <span>
          <strong>${escapeHtml(item.subtitle)}</strong>
          <em>${item.confidence === "needs_review" ? "待复盘" : "有证据"}</em>
        </span>
      </div>
      <ul>${body}</ul>
    </article>`;
  }).join("")}</section>`;
}

function buildDirectoryCounts() {
  const counts = new Map();
  files.forEach(file => {
    const parts = file.path.split("/");
    for (let index = 1; index < parts.length; index += 1) {
      const path = parts.slice(0, index).join("/");
      counts.set(path, (counts.get(path) || 0) + 1);
    }
  });
  return counts;
}

const directoryCounts = buildDirectoryCounts();

function renderTreeButtons(container, counter, currentPath, onSelect) {
  if (!container) return;
  const paths = Array.from(directoryCounts.keys()).sort((a, b) => a.localeCompare(b, "zh-Hans-CN"));
  if (counter) counter.textContent = `${formatNumber(paths.length)} 个目录`;
  const allButton = `<button type="button" data-path="" class="${!currentPath ? "is-active" : ""}" style="--depth:0">全部文件 <span class="muted">${formatNumber(files.length)}</span></button>`;
  container.innerHTML = [allButton].concat(paths.map(path => {
    const parts = path.split("/");
    const label = parts[parts.length - 1];
    const depth = parts.length - 1;
    const active = path === currentPath ? " is-active" : "";
    return `<button type="button" data-path="${escapeHtml(path)}" class="${active}" style="--depth:${depth}">${escapeHtml(label)} <span class="muted">${formatNumber(directoryCounts.get(path))}</span></button>`;
  })).join("");
  container.querySelectorAll("button").forEach(button => {
    button.addEventListener("click", () => onSelect(button.dataset.path || ""));
  });
}

function initHomeFilters() {
  fillSelect(yearFilter, Object.keys(stats.yearCounts || {}), "全部年份");
  fillSelect(tagFilter, Object.keys(stats.tagCounts || {}), "全部主题");
  const risks = Array.from(new Set(files.flatMap(file => file.riskTags || []))).sort();
  fillSelect(riskFilter, risks, "全部风险标签");
}

function fileMatches(file) {
  const query = (searchInput?.value || "").trim().toLowerCase();
  const year = yearFilter?.value || "";
  const tag = tagFilter?.value || "";
  const risk = riskFilter?.value || "";
  const pathMatch = !activePath || file.path.startsWith(activePath + "/");
  const yearMatch = !year || file.year === year;
  const tagMatch = !tag || file.tags.includes(tag);
  const riskMatch = !risk || file.riskTags.includes(risk);
  const haystack = [
    file.path, file.title, file.excerpt, file.content,
    file.tags.join(" "), file.riskTags.join(" "), file.headings.join(" "),
    (file.fiveDimensions || []).map(item => `${item.label} ${item.subtitle} ${(item.items || []).join(" ")}`).join(" ")
  ].join(" ").toLowerCase();
  const queryMatch = !query || haystack.includes(query);
  return pathMatch && yearMatch && tagMatch && riskMatch && queryMatch;
}

function fileRow(file) {
  const selected = file.path === selectedFilePath ? " is-selected" : "";
  const excerpt = truncateText(file.excerpt || "附件或空文件，打开原始文件查看。", 108);
  return `<button class="file-row${selected}" type="button" data-path="${escapeHtml(file.path)}">
    <span class="file-main">
      <span class="file-title"><strong>${escapeHtml(file.title)}</strong>${dimensionMini(file)}${pillList(file.riskTags, true)}</span>
      <span class="file-path">${escapeHtml(file.path)}</span>
      <span class="file-excerpt">${escapeHtml(excerpt)}</span>
    </span>
    <span class="row-meta">
      ${pillList(file.tags)}
      <span class="pill neutral">${escapeHtml(file.year)}</span>
      <span class="pill neutral">${escapeHtml(file.extension || "file")}</span>
      <span class="pill neutral">${formatNumber(file.urlCount)} links</span>
    </span>
  </button>`;
}

function renderHomeTree() {
  renderTreeButtons(directoryTree, treeCount, activePath, path => {
    activePath = path;
    renderHomeTree();
    renderFiles();
  });
}

function renderFiles() {
  if (!fileResults) return;
  const visible = files.filter(fileMatches);
  if (!visible.some(file => file.path === selectedFilePath)) {
    selectedFilePath = visible[0]?.path || "";
  }
  fileResults.innerHTML = visible.map(fileRow).join("");
  if (resultCount) resultCount.textContent = `显示 ${formatNumber(visible.length)} / ${formatNumber(files.length)} 个文件`;
  fileResults.querySelectorAll(".file-row").forEach(row => {
    row.addEventListener("click", () => {
      selectedFilePath = row.dataset.path || "";
      renderFiles();
      const file = fileByPath.get(selectedFilePath);
      if (file) {
        renderDetail(file);
        if (window.innerWidth <= 860) fileDetail?.scrollIntoView({ behavior: "smooth", block: "start" });
      }
    });
  });
  const selected = fileByPath.get(selectedFilePath);
  if (selected) renderDetail(selected);
}

function uniqueSourceLinks(file) {
  const seen = new Set();
  return links.filter(link => {
    if (link.sourcePath !== file.path || seen.has(link.url)) return false;
    seen.add(link.url);
    return true;
  }).slice(0, 36);
}

function renderDetail(file) {
  if (!fileDetail) return;
  const sourceLinks = uniqueSourceLinks(file)
    .map(link => `<li><a class="text-link" href="${escapeHtml(link.url)}" target="_blank" rel="noreferrer">${escapeHtml(link.url)}</a></li>`)
    .join("");
  const content = file.content
    ? `<details class="source-preview"><summary>原文预览</summary><pre>${escapeHtml(file.content)}</pre></details>`
    : `<p class="muted">附件资料可通过公开入口打开或下载。</p>`;
  fileDetail.innerHTML = `<article class="detail-hero">
      <div class="detail-kicker">${dimensionMini(file)}<span>${escapeHtml(file.extension || "file")}</span></div>
      <h2>${escapeHtml(file.title)}</h2>
      <p class="file-path">${escapeHtml(file.path)}</p>
      <div class="meta-row">${pillList(file.tags)}${pillList(file.riskTags, true)}<span class="pill neutral">${escapeHtml(file.year)}</span><span class="pill neutral">${formatBytes(file.size)}</span></div>
      <div class="detail-actions">
        <a class="button primary" href="${fileHref(file)}" target="_blank" rel="noreferrer">打开原始文件</a>
        <a class="button secondary" href="${githubHref(file)}" target="_blank" rel="noreferrer">GitHub 查看</a>
      </div>
    </article>
    <div class="dimension-title">
      <span>道 · 法 · 术 · 器 · 势</span>
      <strong>五维拆解</strong>
    </div>
    ${dimensionCards(file)}
    ${content}
    ${sourceLinks ? `<h3>本文外链</h3><ul class="source-list">${sourceLinks}</ul>` : ""}`;
}

function renderAttachments() {
  if (!attachmentGrid) return;
  if (attachmentCount) attachmentCount.textContent = `${formatNumber(attachments.length)} 个附件`;
  attachmentGrid.innerHTML = attachments.map(file => {
    const detail = file.attachment?.kind === "pptx"
      ? `${file.attachment.slides.length} 页幻灯片，${file.attachment.mediaCount || 0} 个媒体`
      : file.attachment?.kind === "xlsx"
        ? `${file.attachment.sheets?.[0]?.rows || 0} 行，${file.attachment.sheets?.[0]?.columns || 0} 列`
        : "图片附件";
    return `<article class="attachment-card">
      <h3>${escapeHtml(file.title)}</h3>
      <p class="muted">${escapeHtml(file.path)}</p>
      <p>${escapeHtml(detail)}</p>
      <div class="detail-actions"><a class="text-link" href="${fileHref(file)}" target="_blank" rel="noreferrer">打开/下载</a></div>
    </article>`;
  }).join("");
}

function renderDomains() {
  if (!domainGrid) return;
  domainGrid.innerHTML = (stats.topDomains || []).slice(0, 14).map(item =>
    `<article class="domain-card"><strong>${escapeHtml(item.domain)}</strong><span>${formatNumber(item.count)} 次</span></article>`
  ).join("");
}

function initLinkFilters() {
  const domainCounts = new Map();
  uniqueLinks.forEach(link => domainCounts.set(link.domain, (domainCounts.get(link.domain) || 0) + link.count));
  const domains = Array.from(domainCounts.keys()).sort((a, b) => (domainCounts.get(b) - domainCounts.get(a)) || a.localeCompare(b));
  const groups = Array.from(new Set(uniqueLinks.map(link => link.group))).sort();
  fillSelect(domainFilter, domains, "全部域名", domain => `${domain} (${formatNumber(domainCounts.get(domain))})`);
  fillSelect(groupFilter, groups, "全部分组", group => groupLabels[group] || group);
}

function linkMatches(link) {
  const query = (linkSearchInput?.value || "").trim().toLowerCase();
  const domain = domainFilter?.value || "";
  const group = groupFilter?.value || "";
  const domainMatch = !domain || link.domain === domain;
  const groupMatch = !group || link.group === group;
  const haystack = [
    link.url, link.domain, groupLabels[link.group] || link.group, link.tags.join(" "), link.riskTags.join(" "),
    link.sources.map(source => `${source.path} ${source.title}`).join(" ")
  ].join(" ").toLowerCase();
  const queryMatch = !query || haystack.includes(query);
  return domainMatch && groupMatch && queryMatch;
}

function renderLinkDomainChips() {
  if (!linkDomainChips) return;
  const activeDomain = domainFilter?.value || "";
  linkDomainChips.innerHTML = (stats.topDomains || []).slice(0, 24).map(item => {
    const active = item.domain === activeDomain ? " is-active" : "";
    return `<button class="domain-chip${active}" type="button" data-domain="${escapeHtml(item.domain)}">${escapeHtml(item.domain)} · ${formatNumber(item.count)}</button>`;
  }).join("");
  linkDomainChips.querySelectorAll("button").forEach(button => {
    button.addEventListener("click", () => {
      if (domainFilter) domainFilter.value = button.dataset.domain || "";
      linkPage = 1;
      renderLinkDomainChips();
      renderLinks();
    });
  });
}

function renderLinks() {
  if (!linkTableBody) return;
  const visible = uniqueLinks.filter(linkMatches);
  const totalPages = Math.max(1, Math.ceil(visible.length / linkPageSize));
  linkPage = Math.min(Math.max(1, linkPage), totalPages);
  const start = (linkPage - 1) * linkPageSize;
  const pageItems = visible.slice(start, start + linkPageSize);
  if (linkCount) linkCount.textContent = `显示 ${formatNumber(visible.length)} / ${formatNumber(uniqueLinks.length)} 个唯一链接`;
  if (linkPager) linkPager.textContent = `${formatNumber(linkPage)} / ${formatNumber(totalPages)}`;
  if (linkPrev) linkPrev.disabled = linkPage <= 1;
  if (linkNext) linkNext.disabled = linkPage >= totalPages;
  linkTableBody.innerHTML = pageItems.map(link => {
    const first = link.sources[0] || {};
    const sourceFile = fileByPath.get(first.path);
    const sourceText = sourceFile?.excerpt || first.title || first.path || "";
    return `<article class="link-card">
      <div>
        <a class="link-url" href="${escapeHtml(link.url)}" target="_blank" rel="noreferrer">${escapeHtml(link.url)}</a>
        <p class="link-context">${escapeHtml(truncateText(sourceText, 150))}</p>
        <p class="muted">来源：<a class="text-link" href="${rootPrefix}${escapeHtml(first.publicPath || "")}" target="_blank" rel="noreferrer">${escapeHtml(first.path || "unknown")}</a></p>
      </div>
      <div class="link-side">
        ${sourceFile ? dimensionMini(sourceFile) : ""}
        <span>${escapeHtml(link.domain)}</span>
        <span>${escapeHtml(groupLabels[link.group] || link.group)}</span>
        <span>${formatNumber(link.count)} 次出现</span>
        <span class="link-tags">${pillList(link.tags)}${pillList(link.riskTags, true)}</span>
      </div>
    </article>`;
  }).join("");
}

function initRawFilters() {
  const extensions = Array.from(new Set(files.map(file => file.extension || "no-ext"))).sort();
  fillSelect(rawTypeFilter, extensions, "全部类型", value => value === "no-ext" ? "无扩展名" : value);
}

function rawMatches(file) {
  const query = (rawSearchInput?.value || "").trim().toLowerCase();
  const type = rawTypeFilter?.value || "";
  const pathMatch = !rawActivePath || file.path.startsWith(rawActivePath + "/");
  const typeMatch = !type || (file.extension || "no-ext") === type;
  const haystack = [file.path, file.title, file.type, file.year, file.tags.join(" ")].join(" ").toLowerCase();
  const queryMatch = !query || haystack.includes(query);
  return pathMatch && typeMatch && queryMatch;
}

function renderRawTree() {
  renderTreeButtons(rawDirectoryTree, rawTreeCount, rawActivePath, path => {
    rawActivePath = path;
    renderRawTree();
    renderRawTable();
  });
}

function renderRawTable() {
  if (!rawTableBody) return;
  const visible = files.filter(rawMatches);
  if (rawCount) rawCount.textContent = `显示 ${formatNumber(visible.length)} / ${formatNumber(files.length)} 个文件`;
  rawTableBody.innerHTML = visible.map(file => `<tr>
    <td class="raw-path-cell" data-label="文件"><a href="${fileHref(file)}" target="_blank" rel="noreferrer">${escapeHtml(file.path)}</a></td>
    <td data-label="类型">${escapeHtml(file.type)}</td>
    <td data-label="年份">${escapeHtml(file.year)}</td>
    <td data-label="标签">${pillList(file.tags)}${pillList(file.riskTags, true)}</td>
    <td data-label="大小">${formatBytes(file.size)}</td>
  </tr>`).join("");
}

function wireEvents() {
  [searchInput, yearFilter, tagFilter, riskFilter].forEach(control => {
    control?.addEventListener("input", renderFiles);
    control?.addEventListener("change", renderFiles);
  });
  [linkSearchInput, domainFilter, groupFilter].forEach(control => {
    control?.addEventListener("input", () => { linkPage = 1; renderLinks(); });
    control?.addEventListener("change", () => { linkPage = 1; renderLinkDomainChips(); renderLinks(); });
  });
  clearLinkFilters?.addEventListener("click", () => {
    if (linkSearchInput) linkSearchInput.value = "";
    if (domainFilter) domainFilter.value = "";
    if (groupFilter) groupFilter.value = "";
    linkPage = 1;
    renderLinkDomainChips();
    renderLinks();
  });
  linkPrev?.addEventListener("click", () => { linkPage -= 1; renderLinks(); });
  linkNext?.addEventListener("click", () => { linkPage += 1; renderLinks(); });
  [rawSearchInput, rawTypeFilter].forEach(control => {
    control?.addEventListener("input", renderRawTable);
    control?.addEventListener("change", renderRawTable);
  });
}

statCards();
initHomeFilters();
renderHomeTree();
renderFiles();
renderAttachments();
renderDomains();
initLinkFilters();
renderLinkDomainChips();
renderLinks();
initRawFilters();
renderRawTree();
renderRawTable();
wireEvents();
"""
    (SITE_DIR / "app.js").write_text(app, encoding="utf-8")


def write_gitignore() -> None:
    (ROOT / ".gitignore").write_text(".DS_Store\n__pycache__/\n*.pyc\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE, help="MoneyHunter source directory")
    parser.add_argument("--skip-secret-scan", action="store_true", help="Skip high-confidence credential scan")
    args = parser.parse_args()

    source = args.source.expanduser().resolve()
    if not source.exists():
        raise SystemExit(f"Source directory does not exist: {source}")

    if not args.skip_secret_scan:
        findings = scan_for_blocking_secrets(source)
        if findings:
            write_json(DATA_DIR / "blocking-secret-findings.json", findings)
            raise SystemExit(
                "Blocking credential-like secrets were found. "
                "See data/blocking-secret-findings.json before publishing."
            )

    copied = copy_raw_files(source)
    inventory = build_inventory(source, copied)
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    write_json(DATA_DIR / "moneyhunter-manifest.json", inventory)
    write_json(DATA_DIR / "moneyhunter-links.json", inventory["uniqueLinks"])
    write_data_js(inventory)
    generate_docs(inventory)
    generate_site()
    generate_raw_index(inventory)
    generate_readme(inventory)
    write_gitignore()

    print(
        "Imported MoneyHunter: "
        f"{inventory['stats']['fileCount']} files, "
        f"{inventory['stats']['markdownCount']} markdown files, "
        f"{inventory['stats']['uniqueLinkCount']} unique links"
    )


if __name__ == "__main__":
    main()
